import pandas as pd
import openpyxl

path = "C:\\Users\\hrith\\OneDrive\\Documents\\CodeGenerator.xlsm"
file = open('NNF_HEaders_functionsDefination.txt', 'w')
file_for_Test = open('UnitTest_for_NNF_Header.txt', 'w')
file_for_Structures = open('NNF_Header_structures.txt', 'w')
file_for_SEND_RECEIVE_functions =open('NNF_boxconnection_and_login_send_receive_function.txt','w')
file_for_nnf_assembler=open('nnf_assembler.txt','w');
file_for_nnf_transceiver= open('nnf_transceiver_signals.txt','w')
file_for_nnf_switchCase=open ('nnf_switch_case.txt','w')
file_for_host_transceiver_function = open('host_Transceiver.txt','w')
file_for_host_assembler = open('host_assembler.txt','w')
file_for_host_transceiver_signals = open('host_transceiver_signals.txt','w')
file_for_host_switchCase = open('host_switchCase.txt','w')



wb_obj = openpyxl.load_workbook(path)

constructor_declaration_for_bytes = ' '
constructor_declaration_for_value = ' '

for k in range(0, 8 + 6 + 9 + 8+1):  # increment one whenever you add a sheet
    wb_obj.active = k  # change the active sheet starting with 0

    for i in range(1, wb_obj.active.max_column + 1):
        for j in range(1, wb_obj.active.max_row + 1):
            if wb_obj.active.cell(j, i).value is not None:
                if wb_obj.active.cell(j, i).value == "Structure_name":
                    # Constructor taking Byte array
                    file.write(str(wb_obj.active.cell(j + 1, i).value) + "::" + str(
                        wb_obj.active.cell(j + 1, i).value) + "()\n{\n}\n")
                    constructor_declaration_for_bytes = str(wb_obj.active.cell(j + 1, i).value) + "(uint8_t data[]);\n"
                    file.write(str(wb_obj.active.cell(j + 1, i).value) + "::" + str(
                        wb_obj.active.cell(j + 1, i).value) + "(uint8_t data[])\n")
                    file.write("{\n")
                    file.write("*this=*((" + str(wb_obj.active.cell(j + 1, i).value) + "*)data);\n")
                    file.write("}\n")

                    # Constructor taking Parameters
                    parameters = " "
                    default_value = " "
                    for l in range(j + 3, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(l, i + 4).value is None:  # to check default value
                            if wb_obj.active.cell(l, i + 3).value is not None:
                                if (l is not wb_obj.active.max_row):
                                    parameters = parameters + str(wb_obj.active.cell(l, i + 1).value) + " " + str(
                                        wb_obj.active.cell(l, i + 2).value) + "[],"
                                else:
                                    parameters = parameters + str(wb_obj.active.cell(l, i + 1).value) + " " + str(
                                        wb_obj.active.cell(l, i + 2).value) + "[]"

                            else:
                                if (l is not wb_obj.active.max_row):
                                    parameters = parameters + str(wb_obj.active.cell(l, i + 1).value) + " " + str(
                                        wb_obj.active.cell(l, i + 2).value) + ","
                                else:
                                    parameters = parameters + str(wb_obj.active.cell(l, i + 1).value) + " " + str(
                                        wb_obj.active.cell(l, i + 2).value)
                        else:
                            default_value = str(wb_obj.active.cell(l, i + 4).value)
                    constructor_declaration_for_value = str(
                        wb_obj.active.cell(j + 1, i).value) + "(" + parameters + ");\n"
                    file.write(str(wb_obj.active.cell(j + 1, i).value) + "::" + str(
                        wb_obj.active.cell(j + 1, i).value) + "(" + parameters + ")\n")
                    class_name = str(wb_obj.active.cell(j + 1, i).value)
                    file.write("{\n")
                    if wb_obj.active.cell(2, 6).value is not None:
                        trans_code = str(wb_obj.active.cell(2, 6).value)
                        if wb_obj.active.cell(3, 1).value is None:  # to handle internal structures
                            file.write(
                                "this->Direct_Interface_Header.set_values(sizeof(" + class_name + "));\nthis->Message_Header.set_values(NNF_Constants::TransactionCode::" + trans_code)
                            file.write(",(sizeof(" + class_name + ") - sizeof(DIRECT_INTERFACE_HEADER))" + ");\n")
                    else:
                        if wb_obj.active.cell(3, 1).value is None:  # to handle internal structures
                            file.write(
                                "this->Direct_Interface_Header.set_values(sizeof(" + class_name + "));\nthis->Message_Header.set_values(NNF_Constants::TransactionCode::" + class_name)
                            file.write(",(sizeof(" + class_name + ") - sizeof(DIRECT_INTERFACE_HEADER))" + ");\n")
                    for l in range(j + 3, wb_obj.active.max_row + 1):

                        if wb_obj.active.cell(l, i + 3).value is None:
                            if wb_obj.active.cell(l, i + 4).value is None:
                                file.write("this->" + str(wb_obj.active.cell(l, i + 2).value) + " = " + str(
                                    wb_obj.active.cell(l, i + 2).value) + ";\n")
                            else:
                                file.write(
                                    "this->" + str(wb_obj.active.cell(l, i + 2).value) + " = " + default_value + ";\n")
                        else:

                            if wb_obj.active.cell(l, i + 4).value is None:
                                data_type = str(wb_obj.active.cell(l, 2).value)
                                data_type = data_type.replace(" ", "")
                                if data_type == 'uint16_t' or data_type == 'uint8_t' or data_type == 'uint32_t' or data_type == 'uint64_t':
                                    file.write("std::copy(" + str(wb_obj.active.cell(l, i + 2).value) + "," + str(
                                        wb_obj.active.cell(l, i + 2).value) + "+ sizeof(this->" + str(
                                        wb_obj.active.cell(l, i + 2).value) + "), this->" + str(
                                        wb_obj.active.cell(l, i + 2).value) + ");\n")
                                else:# for struct array
                                    itr= wb_obj.active.cell(l, i + 3).value
                                    for m in range(0,itr):
                                        file.write(
                                            "this->" + str(
                                                wb_obj.active.cell(l, i + 2).value) +"["+str(m)+"]"+ " = " + str(wb_obj.active.cell(l, i + 2).value) +"["+str(m)+"]"+ ";\n")

                            else:
                                default_value = str(wb_obj.active.cell(l, i + 4).value)
                                file.write("for(int i=0;i<sizeof(this->" + str(
                                    wb_obj.active.cell(l, i + 2).value) + ");i++)\n{\nthis->" +
                                           str(wb_obj.active.cell(l, i + 2).value) + "[i]=" + default_value + ";\n}\n")
                    file.write("}\n")

                    # Flip_Endian Function
                    file.write("void " + str(wb_obj.active.cell(j + 1, i).value) + "::" + " Flip_endian()\n")
                    file.write("{\n")
                    # if wb_obj.active.cell(3, 1).value is None:  # to handle internal structures
                    # file.write("this->Direct_Interface_Header.Flip_endian();\nthis->Message_Header.Flip_endian();\n")
                    for l in range(j + 1, wb_obj.active.max_row + 1):
                        data_type = str(wb_obj.active.cell(l, 2).value)
                        data_type = data_type.replace(" ", "")
                        if data_type == 'uint16_t' or data_type == 'uint8_t' or data_type == 'uint32_t' or data_type == 'uint64_t':
                            if wb_obj.active.cell(l, i + 3).value is None:
                                file.write("this->" + str(wb_obj.active.cell(l,
                                                                             i + 2).value) + " = " + "boost::endian::endian_reverse(this->" + str(
                                    wb_obj.active.cell(l, i + 2).value) + ");\n")
                            # else:
                            # it is an array so no need to flip

                        else:
                            if wb_obj.active.cell(l,i + 2).value is not None:
                                if wb_obj.active.cell(l, i + 3).value is None and wb_obj.active.cell(l,
                                                                                                     2).value != 'uint8_t':
                                    file.write("this->" + str(wb_obj.active.cell(l,i + 2).value) + ".Flip_endian();\n")
                                else:  # structure array
                                    itr = wb_obj.active.cell(l, i + 3).value
                                    for m in range(0, itr):
                                        file.write("this->" + str(wb_obj.active.cell(l, i + 2).value) + "[" + str(
                                            m) + "]" + ".Flip_endian();\n")

                    file.write("}\n")

                    # toJSON function
                    file.write("std::string " + str(wb_obj.active.cell(j + 1, i).value) + "::" + " toJSON()\n")
                    file.write("{\n")
                    file.write("std::stringstream name;\n json_spirit::Object addr_obj;\n ")

                    doubleQuotes = '"'
                    for l in range(j + 1, wb_obj.active.max_row + 1):
                        data_type = str(wb_obj.active.cell(l, 2).value)
                        data_type=data_type.replace(" ", "")
                        if data_type == 'uint16_t' or data_type == 'uint8_t' or data_type == 'uint32_t' or data_type == 'uint64_t':

                            if wb_obj.active.cell(l, i + 3).value is None:
                                file.write("addr_obj.push_back(json_spirit::Pair(" + doubleQuotes + str(
                                    wb_obj.active.cell(l, i + 2).value) + doubleQuotes + " , " + "this->" + str(
                                    wb_obj.active.cell(l, i + 2).value) + "));\n")
                            else:

                                file.write("addr_obj.push_back(json_spirit::Pair(" + doubleQuotes + str(
                                    wb_obj.active.cell(l, i + 2).value) + doubleQuotes +
                                           " , toString(this->" + str(
                                    wb_obj.active.cell(l, i + 2).value) + ",sizeof(this->" + str(
                                    wb_obj.active.cell(l, i + 2).value) + "))));\n")
                        else:
                            if wb_obj.active.cell(l, i + 2).value is not None:
                                if wb_obj.active.cell(l, i + 3).value is None:
                                    file.write("name<<this->" + str(
                                        wb_obj.active.cell(l, i + 2).value) + ".toJSON();\n")
                                else: # structure array
                                    itr = wb_obj.active.cell(l, i + 3).value
                                    for m in range (0, itr) :
                                        file.write("name<<this->" + str(wb_obj.active.cell(l, i + 2).value) + "["+str(m)+"]"+".toJSON();\n")


                                        # file.write("addr_obj.push_back(json_spirit::Pair(" + doubleQuotes + str(
                                        # wb_obj.active.cell(l, i + 2).value) + doubleQuotes +
                                        #        " , toString(this->" + str(
                                        # wb_obj.active.cell(l, i + 2).value) + ",sizeof(this->" + str(
                                        # wb_obj.active.cell(l, i + 2).value) + "))));\n")

                    file.write("\nwrite(addr_obj, name, json_spirit::pretty_print);\n")
                    file.write("return name.str();\n}\n")
                    # Test for constructor taking byte array
                    file_for_Test.write("TEST(" + class_name + "_Test" + "," + class_name + "_by_Bytes" + ")\n{\n")
                    data = ' '
                    for data_row in range(2, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(data_row, 7).value is not None:
                            if data_row == 2:
                                if wb_obj.active.cell(data_row, 4).value is not None:
                                    for itr in range(1, wb_obj.active.cell(data_row, 4).value + 1):
                                        if itr == 1:
                                            data = str(wb_obj.active.cell(data_row, 7).value)
                                        else:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                else:
                                    data = data + "," + str(wb_obj.active.cell(data_row, 7).value)

                            else:
                                if wb_obj.active.cell(data_row, 4).value is not None:
                                    for itr in range(1, wb_obj.active.cell(data_row, 4).value + 1):
                                        if itr == 1:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                        else:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                else:
                                    data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                    file_for_Test.write("uint8_t data[sizeof(" + class_name + ")]={" + data + "};\n")
                    file_for_Test.write(class_name + " message(data);\n")
                    for l in range(j + 3, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(l, i + 3).value is not None:
                            file_for_Test.write("for (int j = 0; j < sizeof(message." + str(
                                wb_obj.active.cell(l, i + 2).value) + "); j++)\n{\n")
                            file_for_Test.write(
                                "EXPECT_EQ(message." + str(wb_obj.active.cell(l, i + 2).value) + "[j]," + str(
                                    wb_obj.active.cell(l, 9).value) + ");\n}\n")
                        else:
                            file_for_Test.write(
                                "EXPECT_EQ(message." + str(wb_obj.active.cell(l, i + 2).value) + "," + str(
                                    wb_obj.active.cell(l, 9).value) + ");\n")

                    file_for_Test.write("\n}\n")

                    # Test for constructor taking parameters
                    file_for_Test.write("TEST(" + class_name + "_Test" + "," + class_name + "_by_value" + ")\n{\n")
                    data = ' '
                    for data_row in range(2, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(data_row, 7).value is not None:
                            if data_row == 2:
                                if wb_obj.active.cell(data_row, 4).value is not None:
                                    for itr in range(1, wb_obj.active.cell(data_row, 4).value + 1):
                                        if itr == 1:
                                            data = str(wb_obj.active.cell(data_row, 7).value)
                                        else:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                else:
                                    data = data + "," + str(wb_obj.active.cell(data_row, 7).value)

                            else:
                                if wb_obj.active.cell(data_row, 4).value is not None:
                                    for itr in range(1, wb_obj.active.cell(data_row, 4).value + 1):
                                        if itr == 1:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                        else:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                else:
                                    data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                    file_for_Test.write("uint8_t data[sizeof(" + class_name + ")]={" + data + "};\n")
                    file_for_Test.write(class_name + " message0(data);\n")
                    parameters = ' '
                    for l in range(j + 3, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(l, i + 4).value is None:  # to check default value
                            if l is not j + 3:
                                parameters = parameters + ",message0." + str(wb_obj.active.cell(l, i + 2).value)

                            else:
                                parameters = parameters + "message0." + str(wb_obj.active.cell(l, i + 2).value)

                    file_for_Test.write(class_name + " message1(" + parameters + ");\n")
                    for l in range(j + 3, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(l, i + 4).value is None:
                            if wb_obj.active.cell(l, i + 3).value is not None:
                                file_for_Test.write("for (int j = 0; j < sizeof(message1." + str(
                                    wb_obj.active.cell(l, i + 2).value) + "); j++)\n{\n")
                                file_for_Test.write(
                                    "EXPECT_EQ(message1." + str(wb_obj.active.cell(l, i + 2).value) + "[j]," + str(
                                        wb_obj.active.cell(l, 9).value) + ");\n}\n")
                            else:
                                file_for_Test.write(
                                    "EXPECT_EQ(message1." + str(wb_obj.active.cell(l, i + 2).value) + "," + str(
                                        wb_obj.active.cell(l, 9).value) + ");\n")
                        else:
                            if wb_obj.active.cell(l, i + 3).value is not None:
                                file_for_Test.write("for (int j = 0; j < sizeof(message1." + str(
                                    wb_obj.active.cell(l, i + 2).value) + "); j++)\n{\n")
                                file_for_Test.write(
                                    "EXPECT_EQ(message1." + str(wb_obj.active.cell(l, i + 2).value) + "[j]," + str(
                                        wb_obj.active.cell(l, i + 4).value) + ");\n}\n")
                            else:
                                file_for_Test.write(
                                    "EXPECT_EQ(message1." + str(wb_obj.active.cell(l, i + 2).value) + "," + str(
                                        wb_obj.active.cell(l, i + 4).value) + ");\n")
                    file_for_Test.write("\n}\n")
                    # test for Flip endian function
                    file_for_Test.write("TEST(" + class_name + "_Test" + "," + "Flip_endian" + ")\n{\n")
                    data = ' '
                    for data_row in range(2, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(data_row, 7).value is not None:
                            if data_row == 2:
                                if wb_obj.active.cell(data_row, 4).value is not None:
                                    for itr in range(1, wb_obj.active.cell(data_row, 4).value + 1):
                                        if itr == 1:
                                            data = str(wb_obj.active.cell(data_row, 7).value)
                                        else:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                else:
                                    data = data + "," + str(wb_obj.active.cell(data_row, 7).value)

                            else:
                                if wb_obj.active.cell(data_row, 4).value is not None:
                                    for itr in range(1, wb_obj.active.cell(data_row, 4).value + 1):
                                        if itr == 1:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                        else:
                                            data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                                else:
                                    data = data + "," + str(wb_obj.active.cell(data_row, 7).value)
                    file_for_Test.write("uint8_t data[sizeof(" + class_name + ")]={" + data + "};\n")
                    file_for_Test.write(class_name + " message(data);\n")
                    file_for_Test.write("message.Flip_endian();\n")
                    for l in range(j + 3, wb_obj.active.max_row + 1):
                        if wb_obj.active.cell(l, i + 3).value is not None:
                            file_for_Test.write("for (int j = 0; j < sizeof(message." + str(
                                wb_obj.active.cell(l, i + 2).value) + "); j++)\n{\n")
                            file_for_Test.write(
                                "EXPECT_EQ(message." + str(wb_obj.active.cell(l, i + 2).value) + "[j]," + str(
                                    wb_obj.active.cell(l, 8).value) + ");\n}\n")
                        else:
                            file_for_Test.write(
                                "EXPECT_EQ(message." + str(wb_obj.active.cell(l, i + 2).value) + "," + str(
                                    wb_obj.active.cell(l, 8).value) + ");\n")
                    file_for_Test.write("\n}\n")

                    # Structures_Generations
                    if (wb_obj.active.cell(3, 1).value is not None):
                        file_for_Structures.write(
                            "\nstruct\n" + " " + str(wb_obj.active.cell(j + 1, i).value) + "\n{\n")
                        for l in range(j + 3, wb_obj.active.max_row + 1):
                            if (wb_obj.active.cell(l, 4).value is None):
                                if (wb_obj.active.cell(l, 10).value is None):
                                    file_for_Structures.write("  " + str(wb_obj.active.cell(l, 2).value) + " " + str(
                                        wb_obj.active.cell(l, 3).value) + ";\n")
                                else:
                                    file_for_Structures.write("  " + str(wb_obj.active.cell(l, 2).value) + " " + str(
                                        wb_obj.active.cell(l, 3).value) + ":" + str(
                                        wb_obj.active.cell(l, 10).value) + ";\n")

                            elif (wb_obj.active.cell(l, 4).value is not None):
                                file_for_Structures.write("  " + str(wb_obj.active.cell(l, 2).value) + " " + str(
                                    wb_obj.active.cell(l, 3).value) + "[" + str(
                                    wb_obj.active.cell(l, 4).value) + "];\n")
                        file_for_Structures.write(str(wb_obj.active.cell(2, 1).value) + "();\n")
                        file_for_Structures.write(
                            "  " + constructor_declaration_for_bytes + "  " + constructor_declaration_for_value)
                        file_for_Structures.write("  void Flip_endian();\n  std::string toJSON();\n");
                        file_for_Structures.write("\n};")

                    else:
                        file_for_Structures.write(
                            "\nstruct\n" + " " + str(wb_obj.active.cell(j + 1, i).value) + "\n{\n")
                        file_for_Structures.write(
                            "  DIRECT_INTERFACE_HEADER Direct_Interface_Header;\n  MESSAGE_HEADER Message_Header;\n")
                        for l in range(4, wb_obj.active.max_row + 1):
                            if (wb_obj.active.cell(l, 4).value is None):
                                if (wb_obj.active.cell(l, 10).value is None):
                                    file_for_Structures.write("  " + str(wb_obj.active.cell(l, 2).value) + " " + str(
                                        wb_obj.active.cell(l, 3).value) + ";\n")
                                else:
                                    file_for_Structures.write("  " + str(wb_obj.active.cell(l, 2).value) + " " + str(
                                        wb_obj.active.cell(l, 3).value) + ":" + str(
                                        wb_obj.active.cell(l, 10).value) + ";\n")


                            elif (wb_obj.active.cell(l, 4).value is not None):
                                file_for_Structures.write("  " + str(wb_obj.active.cell(l, 2).value) + " " + str(
                                    wb_obj.active.cell(l, 3).value) + "[" + str(
                                    wb_obj.active.cell(l, 4).value) + "];\n")
                        file_for_Structures.write(
                            "  " + constructor_declaration_for_bytes + "  " + constructor_declaration_for_value)
                        file_for_Structures.write("  void Flip_endian();\n  std::string toJSON();\n");
                        file_for_Structures.write("\n};")
file.close()
file_for_Test.close()
file_for_Structures.close()
# sheet_obj = wb_obj.active
# cell_obj = sheet_obj.cell(row=2, column=1)
# struct_name = cell_obj.value
# print("struct " + struct_name + "{\n};")
for k in range(0, 8 + 6 + 9 + 8+1):  # increment one whenever you add a sheet
    wb_obj.active = k  # change the active sheet starting with 0
    structure_name = str(wb_obj.active.cell(2, 1).value)
    newLine = "\\" + 'n'
# NNF End
    if wb_obj.active.cell(2,10).value == 'NNF' and wb_obj.active.cell(3,10).value == 'NSE':
        #send function
        file_for_SEND_RECEIVE_functions.write('void NNF_logIN::SEND_'+structure_name+'()\n{\nBOOST_LOG_TRIVIAL(info) << "SEND_'+structure_name+'L: " << __LINE__ << ":" << __FILE__ "'+newLine+'";\n')
        elements = ' '
        for i in range(2, wb_obj.active.max_row + 1):

            if wb_obj.active.cell(i, 7).value is not None:
                if i == 2:
                    elements = elements + str(wb_obj.active.cell(i, 7).value)
                else:
                    elements = elements + str(wb_obj.active.cell(i, 7).value) + ','
        file_for_SEND_RECEIVE_functions.write("uint8_t data[sizeof(MS_UPDATE_LOCAL_DATABASE)] = { " + elements + "};\n")
        file_for_SEND_RECEIVE_functions.write(structure_name+'* ms = new '+structure_name+'(data);\n')
        file_for_SEND_RECEIVE_functions.write("ms->Direct_Interface_Header.Length = sizeof("+structure_name+");\n")
        transeCode=str(wb_obj.active.cell(2,6).value)
        file_for_SEND_RECEIVE_functions.write("ms->Message_Header.TransactionCode = NNF_Constants::TransactionCode::"+ transeCode+";\n")
        file_for_SEND_RECEIVE_functions.write("ms->Message_Header.MessageLength = (sizeof("+structure_name+") - sizeof(DIRECT_INTERFACE_HEADER));")
        file_for_SEND_RECEIVE_functions.write('BOOST_LOG_TRIVIAL(info) << ms->toJSON() << " L: " << __LINE__ << ":" << __FILE__ << "'+ newLine+'";\n')
        file_for_SEND_RECEIVE_functions.write('ms->Flip_endian();\n')
        file_for_SEND_RECEIVE_functions.write('sig_SEND((uint8_t*)ms, sizeof('+structure_name+'));\n')
        file_for_SEND_RECEIVE_functions.write("}\n")
        # the receive function for nnf
        file_for_SEND_RECEIVE_functions.write(
            'void NNF_logIN::RECEIVE_' + structure_name + '(' + structure_name + '* ms)\n{\n')
        file_for_SEND_RECEIVE_functions.write(
            'BOOST_LOG_TRIVIAL(info) << "RECEIVE_' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        file_for_SEND_RECEIVE_functions.write(
            'BOOST_LOG_TRIVIAL(info) << ms->toJSON() <<" L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        file_for_SEND_RECEIVE_functions.write('delete  ms;\n}\n')
        # nnf_switch case
        if wb_obj.active.cell(2, 6).value is not None:
            file_for_nnf_switchCase.write(
                'case NNF_Constants::TransactionCode::' + str(wb_obj.active.cell(2, 6).value) + ' :\n{\n')
            file_for_nnf_switchCase.write(
                'BOOST_LOG_TRIVIAL(info) << "Classify_and_Distribute :' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
            file_for_nnf_switchCase.write(structure_name + '* ms=new ' + structure_name + '(data);\n')
            file_for_nnf_switchCase.write('ms->Flip_endian();\n')
            file_for_nnf_switchCase.write('sig_RECEIVE_' + structure_name + '(ms);\n')
            file_for_nnf_switchCase.write('break;\n}\n')
        # nnf_transceiver
        file_for_nnf_transceiver.write(
            'boost::signals2::signal<void(' + structure_name + '*)> sig_RECEIVE_' + structure_name + ';\n')
        # nnf_Assembler
        file_for_nnf_assembler.write(
            'Transciver1.sig_RECEIVE_' + structure_name + '.connect(boost::bind(&NNF_logIN::RECEIVE_' + structure_name + ', &nnf_logIn, _1));\n')

    if wb_obj.active.cell(2,10).value == 'NNF' and wb_obj.active.cell(3,10).value is None:
        # send function

        file_for_SEND_RECEIVE_functions.write(
            'void NNF_logIN::SEND_' + structure_name + '()\n{\nBOOST_LOG_TRIVIAL(info) << "SEND_' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        elements = ' '
        for i in range(2, wb_obj.active.max_row + 1):

            if wb_obj.active.cell(i, 7).value is not None:
                if i == 2:
                    elements = elements + str(wb_obj.active.cell(i, 7).value)
                else:
                    elements = elements + str(wb_obj.active.cell(i, 7).value) + ','
        file_for_SEND_RECEIVE_functions.write("uint8_t data[sizeof(MS_UPDATE_LOCAL_DATABASE)] = { " + elements + "};\n")
        file_for_SEND_RECEIVE_functions.write(structure_name + '* ms = new ' + structure_name + '(data);\n')
        file_for_SEND_RECEIVE_functions.write("ms->Direct_Interface_Header.Length = sizeof(" + structure_name + ");\n")
        transeCode = str(wb_obj.active.cell(2, 6).value)
        file_for_SEND_RECEIVE_functions.write(
            "ms->Message_Header.TransactionCode = NNF_Constants::TransactionCode::" + transeCode + ";\n")
        file_for_SEND_RECEIVE_functions.write(
            "ms->Message_Header.MessageLength = (sizeof(" + structure_name + ") - sizeof(DIRECT_INTERFACE_HEADER));")
        file_for_SEND_RECEIVE_functions.write(
            'BOOST_LOG_TRIVIAL(info) << ms->toJSON() << " L: " << __LINE__ << ":" << __FILE__ << "' + newLine + '";\n')
        file_for_SEND_RECEIVE_functions.write('ms->Flip_endian();\n')
        file_for_SEND_RECEIVE_functions.write('sig_SEND((uint8_t*)ms, sizeof(' + structure_name + '));\n')
        file_for_SEND_RECEIVE_functions.write("}\n")
    if wb_obj.active.cell(2,10).value == 'NSE' and wb_obj.active.cell(3,10).value is None:

        # the receive function for nnf
        file_for_SEND_RECEIVE_functions.write(
            'void NNF_logIN::RECEIVE_' + structure_name + '(' + structure_name + '* ms)\n{\n')
        file_for_SEND_RECEIVE_functions.write(
            'BOOST_LOG_TRIVIAL(info) << "RECEIVE_' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        file_for_SEND_RECEIVE_functions.write(
            'BOOST_LOG_TRIVIAL(info) << ms->toJSON() << "L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        file_for_SEND_RECEIVE_functions.write('delete  ms;\n}\n')
        # nnf_switch case
        if wb_obj.active.cell(2,6).value is not None:
          file_for_nnf_switchCase.write('case NNF_Constants::TransactionCode::'+str(wb_obj.active.cell(2,6).value)+' :\n{\n')
          file_for_nnf_switchCase.write(
              'BOOST_LOG_TRIVIAL(info) << "Classify_and_Distribute :' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
          file_for_nnf_switchCase.write(structure_name+'* ms=new '+structure_name+'(data);\n')
          file_for_nnf_switchCase.write('ms->Flip_endian();\n')
          file_for_nnf_switchCase.write('sig_RECEIVE_'+structure_name+'(ms);\n')
          file_for_nnf_switchCase.write('break;\n}\n')

       #nnf_transceiver
        file_for_nnf_transceiver.write('boost::signals2::signal<void('+structure_name+'*)> sig_RECEIVE_'+structure_name+';\n')
       #nnf_Assembler
        file_for_nnf_assembler.write('Transciver1.sig_RECEIVE_'+structure_name+'.connect(boost::bind(&NNF_logIN::RECEIVE_'+structure_name+', &nnf_logIn, _1));\n')

# HOST_end---------------------------------------------------------------------------------------------------------------------------------------
    if wb_obj.active.cell(2,10).value == 'NNF' and wb_obj.active.cell(3,10).value == 'NSE':
        #send function
        file_for_host_transceiver_function.write('void HOST_logIN::SEND_'+structure_name+'()\n{\nBOOST_LOG_TRIVIAL(info) << "SEND_'+structure_name+'L: " << __LINE__ << ":" << __FILE__ "'+newLine+'";\n')
        elements = ' '
        for i in range(2, wb_obj.active.max_row + 1):

            if wb_obj.active.cell(i, 7).value is not None:
                if i == 2:
                    elements = elements + str(wb_obj.active.cell(i, 7).value)
                else:
                    elements = elements + str(wb_obj.active.cell(i, 7).value) + ','
        file_for_host_transceiver_function.write("uint8_t data[sizeof(MS_UPDATE_LOCAL_DATABASE)] = { " + elements + "};\n")
        file_for_host_transceiver_function.write(structure_name+'* ms = new '+structure_name+'(data);\n')
        file_for_host_transceiver_function.write("ms->Direct_Interface_Header.Length = sizeof("+structure_name+");\n")
        transeCode=str(wb_obj.active.cell(2,6).value)
        file_for_host_transceiver_function.write("ms->Message_Header.TransactionCode = NNF_Constants::TransactionCode::"+ transeCode+";\n")
        file_for_host_transceiver_function.write("ms->Message_Header.MessageLength = (sizeof("+structure_name+") - sizeof(DIRECT_INTERFACE_HEADER));")
        file_for_host_transceiver_function.write('BOOST_LOG_TRIVIAL(info) << ms->toJSON() << " L: " << __LINE__ << ":" << __FILE__ << "'+ newLine+'";\n')
        file_for_host_transceiver_function.write('ms->Flip_endian();\n')
        file_for_host_transceiver_function.write('sig_SEND((uint8_t*)ms, sizeof('+structure_name+'));\n')
        file_for_host_transceiver_function.write("}\n")
        # the receive function for host
        file_for_host_transceiver_function.write(
            'void HOST_logIN::RECEIVE_' + structure_name + '(' + structure_name + '* ms)\n{\n')
        file_for_host_transceiver_function.write(
            'BOOST_LOG_TRIVIAL(info) << "RECEIVE_' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        file_for_host_transceiver_function.write(
            'BOOST_LOG_TRIVIAL(info) << ms->toJSON() << "L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        file_for_host_transceiver_function.write('delete  ms;\n}\n')
        # host_switch case
        if wb_obj.active.cell(2, 6).value is not None:
            file_for_host_switchCase.write(
                'case NNF_Constants::TransactionCode::' + str(wb_obj.active.cell(2, 6).value) + ' :\n{\n')
            file_for_host_switchCase.write(
                'BOOST_LOG_TRIVIAL(info) << "Classify_and_Distribute :' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
            file_for_host_switchCase.write(structure_name + '* ms=new ' + structure_name + '(data);\n')
            file_for_host_switchCase.write('ms->Flip_endian();\n')
            file_for_host_switchCase.write('sig_RECEIVE_' + structure_name + '(ms);\n')
            file_for_host_switchCase.write('break;\n}\n')
        # host_transceiver
        file_for_host_transceiver_signals.write(
            'boost::signals2::signal<void(' + structure_name + '*)> sig_RECEIVE_' + structure_name + ';\n')
        # host_Assembler
        file_for_host_assembler.write(
            'THost_Transceiver01.sig_RECEIVE_' + structure_name + '.connect(boost::bind(&HOST_logIN::RECEIVE_' + structure_name + ', &host_logIn, _1));\n')

    if wb_obj.active.cell(2,10).value == 'NSE' and wb_obj.active.cell(3,10).value is None:
        # send function
        file_for_host_transceiver_function.write(
            'void HOST_logIN::SEND_' + structure_name + '()\n{\nBOOST_LOG_TRIVIAL(info) << "SEND_' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
        elements = ' '
        for i in range(2, wb_obj.active.max_row + 1):

            if wb_obj.active.cell(i, 7).value is not None:
                if i == 2:
                    elements = elements + str(wb_obj.active.cell(i, 7).value)
                else:
                    elements = elements + str(wb_obj.active.cell(i, 7).value) + ','
        file_for_host_transceiver_function.write("uint8_t data[sizeof(MS_UPDATE_LOCAL_DATABASE)] = { " + elements + "};\n")
        file_for_host_transceiver_function.write(structure_name + '* ms = new ' + structure_name + '(data);\n')
        file_for_host_transceiver_function.write("ms->Direct_Interface_Header.Length = sizeof(" + structure_name + ");\n")
        transeCode = str(wb_obj.active.cell(2, 6).value)
        file_for_host_transceiver_function.write("ms->Message_Header.TransactionCode = NNF_Constants::TransactionCode::" + transeCode + ";\n")
        file_for_host_transceiver_function.write("ms->Message_Header.MessageLength = (sizeof(" + structure_name + ") - sizeof(DIRECT_INTERFACE_HEADER));")
        file_for_host_transceiver_function.write('BOOST_LOG_TRIVIAL(info) << ms->toJSON() << " L: " << __LINE__ << ":" << __FILE__ << "' + newLine + '";\n')
        file_for_host_transceiver_function.write('ms->Flip_endian();\n')
        file_for_host_transceiver_function.write('sig_SEND((uint8_t*)ms, sizeof(' + structure_name + '));\n')
        file_for_host_transceiver_function.write("}\n")

    if wb_obj.active.cell(2,10).value == 'NNF' and wb_obj.active.cell(3,10).value is None:

            # the receive function for host
            file_for_host_transceiver_function.write(
                'void HOST_logIN::RECEIVE_' + structure_name + '(' + structure_name + '* ms)\n{\n')
            file_for_host_transceiver_function.write(
                'BOOST_LOG_TRIVIAL(info) << "RECEIVE_' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
            file_for_host_transceiver_function.write(
                'BOOST_LOG_TRIVIAL(info) << ms->toJSON() << "L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
            file_for_host_transceiver_function.write('delete  ms;\n}\n')
        # host_switch case
            if wb_obj.active.cell(2, 6).value is not None:
                file_for_host_switchCase.write(
                    'case NNF_Constants::TransactionCode::' + str(wb_obj.active.cell(2, 6).value) + ' :\n{\n')
                file_for_host_switchCase.write(
                    'BOOST_LOG_TRIVIAL(info) << "Classify_and_Distribute :' + structure_name + 'L: " << __LINE__ << ":" << __FILE__ "' + newLine + '";\n')
                file_for_host_switchCase.write(structure_name + '* ms=new ' + structure_name + '(data);\n')
                file_for_host_switchCase.write('ms->Flip_endian();\n')
                file_for_host_switchCase.write('sig_RECEIVE_' + structure_name + '(ms);\n')
                file_for_host_switchCase.write('break;\n}\n')

                # host_transceiver
                file_for_host_transceiver_signals.write(
                    'boost::signals2::signal<void(' + structure_name + '*)> sig_RECEIVE_' + structure_name + ';\n')
                # host_Assembler
                file_for_host_assembler.write(
                    'THost_Transceiver01.sig_RECEIVE_' + structure_name + '.connect(boost::bind(&HOST_logIN::RECEIVE_' + structure_name + ', &host_logIn, _1));\n')
