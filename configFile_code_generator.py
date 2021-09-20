import openpyxl

path = "config_.xlsm"
file1 = open('config_code\\config_.ini', 'w')
file2 = open('config_code\\NNF_programOptions.txt', 'w')
wb_obj = openpyxl.load_workbook(path)

# make ini file
for i in range(1,wb_obj.active.max_row + 1):
    if wb_obj.active.cell(i,2).value is None:
         file1.write('['+str(wb_obj.active.cell(i,1).value)+']\n')
    else:
         file1.write(str(wb_obj.active.cell(i,1).value)+'='+str(wb_obj.active.cell(i,2).value)+'\n')

file1.close()
# write program options
file2.write(' boost::program_options::options_description config("Configuration");\nconfig.add_options()\n("help", "produce help message")\n("Port,p", boost::program_options::value<int>(&Port)->default_value(0), "Port")\n("IP", boost::program_options::value<std::string>(&IP)->default_value(" "), "IP")\n("File", boost::program_options::value<std::string>(&File)->default_value(" "), "config file with extension")')
double_inverted_commas='"'
seg=''
for i in range(1,wb_obj.active.max_row + 1):
    if wb_obj.active.cell(i,2).value is None:
         seg=str(wb_obj.active.cell(i,1).value)
    else:
        file2.write('('+double_inverted_commas+seg+'.'+str(wb_obj.active.cell(i,1).value)+double_inverted_commas+','+'boost::program_options::value<std::string>()'+')\n')
